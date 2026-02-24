from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from .models import Prospect
from .forms import ProspectForm
import csv
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.core.paginator import Paginator
from django.core.exceptions import PermissionDenied
from .email_utils import envoyer_email_bienvenue, envoyer_email_relance 
from .email_utils import envoyer_email_bienvenue, envoyer_email_relance, notifier_admin_nouveau_prospect

@login_required
def dashboard(request):
    total_prospects = Prospect.objects.count()
    total_clients   = Prospect.objects.filter(statut='client').count()
    nouveaux        = Prospect.objects.filter(statut='nouveau').count()
    perdus          = Prospect.objects.filter(statut='perdu').count()

    taux = round((total_clients / total_prospects * 100), 1) if total_prospects > 0 else 0

    par_statut = Prospect.objects.values('statut').annotate(total=Count('id'))
    derniers   = Prospect.objects.select_related('cree_par').order_by('-date_creation')[:8]

    context = {
        'total_prospects': total_prospects,
        'total_clients':   total_clients,
        'nouveaux':        nouveaux,
        'perdus':          perdus,
        'taux':            taux,
        'par_statut':      par_statut,
        'derniers':        derniers,
    }
    return render(request, 'crm/dashboard.html', context)
   


@login_required
def prospect_list(request):
    prospects = Prospect.objects.all()

    # Recherche
    q = request.GET.get('q')
    if q:
        prospects = prospects.filter(
            Q(nom_entreprise__icontains=q) |
            Q(email__icontains=q) |
            Q(ville__icontains=q)
        )

    # Filtre par statut
    statut = request.GET.get('statut')
    if statut:
        prospects = prospects.filter(statut=statut)

    # Si commercial, il voit seulement ses prospects
    if not request.user.is_superuser and not request.user.groups.filter(name='Manager').exists():
        prospects = prospects.filter(cree_par=request.user)

    context = {
        'prospects':       prospects,
        'statut_choices':  Prospect.STATUT_CHOICES,
        'statut_filtre':   statut or '',
        'q':               q or '',
    }
    return render(request, 'crm/prospect_list.html', context)


@login_required
def prospect_detail(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)
    return render(request, 'crm/prospect_detail.html', {'prospect': prospect})




@login_required
def prospect_create(request):
    if request.method == 'POST':
        form = ProspectForm(request.POST)
        if form.is_valid():
            prospect = form.save(commit=False)
            prospect.cree_par = request.user
            prospect.save()
            # Email au prospect
            envoyer_email_bienvenue(prospect)
            # Notification à toi
            notifier_admin_nouveau_prospect(prospect)
            messages.success(request, '✅ Prospect ajouté avec succès !')
            return redirect('crm:prospect_list')
    else:
        form = ProspectForm()
    return render(request, 'crm/prospect_form.html', {'form': form, 'titre': 'Ajouter un prospect'})





@login_required
def prospect_update(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)
    if request.method == 'POST':
        form = ProspectForm(request.POST, instance=prospect)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Prospect modifié avec succès !')
            return redirect('crm:prospect_list')
    else:
        form = ProspectForm(instance=prospect)
    return render(request, 'crm/prospect_form.html', {'form': form, 'titre': 'Modifier le prospect'})


@login_required
def prospect_delete(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)
    if request.method == 'POST':
        prospect.delete()
        messages.success(request, '🗑️ Prospect supprimé.')
        return redirect('crm:prospect_list')
    return render(request, 'crm/prospect_confirm_delete.html', {'prospect': prospect})@login_required
def dashboard(request):
    total_prospects = Prospect.objects.count()
    total_clients   = Prospect.objects.filter(statut='client').count()
    nouveaux        = Prospect.objects.filter(statut='nouveau').count()
    perdus          = Prospect.objects.filter(statut='perdu').count()

    taux = round((total_clients / total_prospects * 100), 1) if total_prospects > 0 else 0

    par_statut      = Prospect.objects.values('statut').annotate(total=Count('id'))
    derniers        = Prospect.objects.select_related('cree_par').order_by('-date_creation')[:8]

    context = {
        'total_prospects': total_prospects,
        'total_clients':   total_clients,
        'nouveaux':        nouveaux,
        'perdus':          perdus,
        'taux':            taux,
        'par_statut':      par_statut,
        'derniers':        derniers,
    }
    return render(request, 'crm/dashboard.html', context)


@login_required
def export_excel(request):
    prospects = Prospect.objects.select_related('cree_par').all()

    # Si commercial, exporte seulement ses prospects
    if not request.user.is_superuser and not request.user.groups.filter(name='Manager').exists():
        prospects = prospects.filter(cree_par=request.user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # ── Style entête ──────────────────────────────────────
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1a1a2e")
    header_align   = Alignment(horizontal="center", vertical="center")

    headers = [
        'ID', 'Entreprise', 'Email', 'Téléphone',
        'Ville', 'Adresse', 'Secteur', 'Statut',
        'Créé par', 'Date création'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align

    # ── Largeurs colonnes ─────────────────────────────────
    column_widths = [6, 30, 30, 15, 20, 30, 15, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ── Données ───────────────────────────────────────────
    for row_num, p in enumerate(prospects, 2):
        ws.append([
            p.pk,
            p.nom_entreprise,
            p.email        or '',
            p.telephone    or '',
            p.ville        or '',
            p.adresse      or '',
            p.get_secteur_display(),
            p.get_statut_display(),
            p.cree_par.username if p.cree_par else '',
            p.date_creation.strftime('%d/%m/%Y %H:%M'),
        ])
        # Lignes alternées
        if row_num % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor="F0F4FF")

    # ── Réponse HTTP ──────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="prospects_bizcrm.xlsx"'
    wb.save(response)
    return response


@login_required
def import_csv(request):
    if request.method == 'POST':
        fichier = request.FILES.get('fichier_csv')

        if not fichier:
            messages.error(request, '❌ Aucun fichier sélectionné.')
            return redirect('crm:import_csv')

        if not fichier.name.endswith('.csv'):
            messages.error(request, '❌ Le fichier doit être au format CSV.')
            return redirect('crm:import_csv')

        try:
            decoded   = fichier.read().decode('utf-8-sig')
            reader    = csv.DictReader(io.StringIO(decoded))
            success   = 0
            errors    = []

            for i, row in enumerate(reader, start=2):
                nom = row.get('nom_entreprise', '').strip()
                if not nom:
                    errors.append(f'Ligne {i} : nom_entreprise manquant.')
                    continue

                Prospect.objects.create(
                    nom_entreprise = nom,
                    email          = row.get('email',     '').strip() or None,
                    telephone      = row.get('telephone', '').strip() or None,
                    ville          = row.get('ville',     '').strip() or None,
                    adresse        = row.get('adresse',   '').strip() or None,
                    secteur        = row.get('secteur',   'autre').strip(),
                    statut         = row.get('statut',    'nouveau').strip(),
                    notes          = row.get('notes',     '').strip() or None,
                    cree_par       = request.user,
                )
                success += 1

            if success:
                messages.success(request, f'✅ {success} prospect(s) importé(s) avec succès !')
            for err in errors:
                messages.warning(request, f'⚠️ {err}')

        except Exception as e:
            messages.error(request, f'❌ Erreur lors de la lecture du fichier : {e}')

        return redirect('crm:prospect_list')

    return render(request, 'crm/import_csv.html')

@login_required
def telecharger_modele_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="modele_prospects.csv"'

    writer = csv.writer(response)
    writer.writerow(['nom_entreprise', 'email', 'telephone', 'ville', 'adresse', 'secteur', 'statut', 'notes'])
    writer.writerow(['Acme Corp', 'contact@acme.com', '+242 06 000 0000', 'Brazzaville', 'Avenue Foch', 'commerce', 'nouveau', 'Premier contact'])
    writer.writerow(['Beta SARL', 'info@beta.com', '+242 05 111 1111', 'Pointe-Noire', 'Rue Loango', 'services', 'contacte', ''])

    return response

@login_required
def prospect_list(request):
    prospects = Prospect.objects.select_related('cree_par').all()

    # Recherche
    q = request.GET.get('q', '')
    if q:
        prospects = prospects.filter(
            Q(nom_entreprise__icontains=q) |
            Q(email__icontains=q)          |
            Q(ville__icontains=q)
        )

    # Filtre par statut
    statut = request.GET.get('statut', '')
    if statut:
        prospects = prospects.filter(statut=statut)

    # Filtre par secteur
    secteur = request.GET.get('secteur', '')
    if secteur:
        prospects = prospects.filter(secteur=secteur)

    # Si commercial, il voit seulement ses prospects
    if not request.user.is_superuser and not request.user.groups.filter(name='Manager').exists():
        prospects = prospects.filter(cree_par=request.user)

    # Pagination — 10 prospects par page
    paginator = Paginator(prospects, 10)
    page      = request.GET.get('page', 1)
    prospects_page = paginator.get_page(page)

    context = {
        'prospects':       prospects_page,
        'statut_choices':  Prospect.STATUT_CHOICES,
        'secteur_choices': Prospect.SECTEUR_CHOICES,
        'statut_filtre':   statut,
        'secteur_filtre':  secteur,
        'q':               q,
        'total':           paginator.count,
    }
    return render(request, 'crm/prospect_list.html', context)

def check_prospect_access(user, prospect):
    """Vérifie qu'un commercial ne peut accéder qu'à ses propres prospects."""
    if user.is_superuser or user.groups.filter(name='Manager').exists():
        return True
    if prospect.cree_par == user:
        return True
    raise PermissionDenied

@login_required
def prospect_detail(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)
    check_prospect_access(request.user, prospect)
    return render(request, 'crm/prospect_detail.html', {'prospect': prospect})


@login_required
def prospect_update(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)
    check_prospect_access(request.user, prospect)
    if request.method == 'POST':
        form = ProspectForm(request.POST, instance=prospect)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Prospect modifié avec succès !')
            return redirect('crm:prospect_list')
    else:
        form = ProspectForm(instance=prospect)
    return render(request, 'crm/prospect_form.html', {'form': form, 'titre': 'Modifier le prospect'})


@login_required
def prospect_delete(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)
    check_prospect_access(request.user, prospect)
    # Seul un Manager ou Admin peut supprimer
    if not request.user.is_superuser and not request.user.groups.filter(name='Manager').exists():
        raise PermissionDenied
    if request.method == 'POST':
        prospect.delete()
        messages.success(request, '🗑️ Prospect supprimé.')
        return redirect('crm:prospect_list')
    return render(request, 'crm/prospect_confirm_delete.html', {'prospect': prospect})
@login_required
def envoyer_relance(request, pk):
    prospect = get_object_or_404(Prospect, pk=pk)

    if request.method == 'POST':
        message_perso = request.POST.get('message', '')
        success, msg  = envoyer_email_relance(prospect, message_perso)
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
        return redirect('crm:prospect_detail', pk=pk)

    return render(request, 'crm/envoyer_relance.html', {'prospect': prospect})